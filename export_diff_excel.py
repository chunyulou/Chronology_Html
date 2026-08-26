# -*- coding: utf-8 -*-
"""輸出正覺大事紀的分類彙整活頁簿（含現場操作分類欄）。

工作表：
  1. 總表          196 則全部
  2. 差異表        只收錄各案與 1150810 定案不同的則次
  3~9. 各分頁籤     每個分頁籤一張，列出該分類收錄了哪些大事
  最後 說明        本檔的產製規則與資料來源

「現場操作分類」＝**最大化**綜合：把所有提案（含定案）提出過的分頁籤全部列上，
依分頁籤原始順序排列。各分頁籤工作表即依此欄歸類。

各菩薩的說明文字直接寫在該菩薩的分類儲存格內（同一格分行呈現），比照網頁。

用法：
  python export_diff_excel.py [輸出資料夾]
"""
import json
import os
import re
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.abspath(__file__))
DEFAULT_OUT = os.path.join(
    os.path.expanduser('~'),
    'OneDrive', '正覺', '大事紀', '04編寫', '14分頁籤', '第二次分類')
OUT_NAME = '正覺大事紀_分類彙整_含現場操作.xlsx'

HEADER_FILL = PatternFill('solid', fgColor='793C01')
TAB_HEADER_FILL = PatternFill('solid', fgColor='1E7A5A')
DIFF_FILL = PatternFill('solid', fgColor='FDECC8')
BASE_FILL = PatternFill('solid', fgColor='E8F3EA')
LIVE_FILL = PatternFill('solid', fgColor='DCEFE4')
THIN = Side(style='thin', color='D9CBBD')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TOP_WRAP = Alignment(vertical='top', wrap_text=True)


def read_js(path, marker):
    """從 js 檔中取出指定變數的 JSON 陣列／物件。"""
    src = open(os.path.join(BASE, path), encoding='utf-8').read()
    i = src.index(marker) + len(marker)
    opener = src[i]
    closer = {'[': ']', '{': '}'}[opener]
    depth, in_str, esc = 0, False, False
    for j in range(i, len(src)):
        ch = src[j]
        if in_str:
            if esc:
                esc = False
            elif ch == chr(92):
                esc = True
            elif ch == '"':
                in_str = False
            continue
        if ch == '"':
            in_str = True
        elif ch == opener:
            depth += 1
        elif ch == closer:
            depth -= 1
            if depth == 0:
                return json.loads(src[i:j + 1])
    raise ValueError('無法解析 %s 中的 %s' % (path, marker))


class Data(object):
    def __init__(self):
        self.events = read_js('table_data.js', 'const tableData = ')
        self.proposals = read_js('meeting2_data.js', 'const proposals = ')
        self.categories = read_js('meeting2_data.js', 'const eventProposalCategories = ')
        self.notes = read_js('meeting2_data.js', 'const eventProposalNotes = ')
        self.aliases = read_js('meeting2_data.js', 'const tagAliases = ')
        self.tab_order = read_js('meeting2_data.js', 'const TAB_ORDER = ')
        self.baseline = 'D0'
        self.others = [p for p in self.proposals if p['id'] != self.baseline]

    def tags(self, sid, pid):
        return (self.categories.get(str(sid), {}) or {}).get(pid, [])

    def canon(self, tags):
        out = []
        for t in tags:
            c = self.aliases.get(t, t)
            if c not in out:
                out.append(c)
        return out

    def key(self, tags):
        return '｜'.join(sorted(self.canon(tags)))

    def suggesters(self, sid):
        base = self.key(self.tags(sid, self.baseline))
        return [p['id'] for p in self.others
                if self.key(self.tags(sid, p['id'])) != base]

    def live_tags(self, sid):
        """現場操作分類＝所有提案提出過的分頁籤之最大化聯集。"""
        pool = []
        for p in self.proposals:
            for t in self.canon(self.tags(sid, p['id'])):
                if t not in pool:
                    pool.append(t)
        ordered = [t for t in self.tab_order if t in pool]
        ordered += [t for t in pool if t not in ordered]
        return ordered

    def note_raw(self, sid, pid):
        return ((self.notes.get(str(sid), {}) or {}).get(pid) or {}).get('raw', '')

    def cell_text(self, sid, pid):
        """該案的分類＋說明，比照網頁寫在同一格內。"""
        tags = '、'.join(self.tags(sid, pid))
        raw = self.note_raw(sid, pid)
        if not raw:
            return tags
        lines = [ln for ln in raw.replace('\r\n', '\n').split('\n')]
        # 原文第一行若就是分類本身，不重複顯示
        if lines and re.sub(r'[、,，\s]', '', lines[0]) == re.sub(r'[、,，\s]', '', tags):
            lines = lines[1:]
        rest = '\n'.join(ln for ln in lines if ln.strip())
        return tags + ('\n' + rest if rest else '')


def style_header(ws, fill=HEADER_FILL, height=30):
    for c in ws[1]:
        c.fill = fill
        c.font = Font(color='FFFFFF', bold=True)
        c.alignment = Alignment(vertical='center', wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[1].height = height


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_main_sheet(ws, d, rows, title_note):
    header = ['序號', '日期', '大事', '紀要', '現場操作分類（最大化綜合）', '1150810 定案初版']
    header += [p['name'] for p in d.others]
    header += ['提出調整人數', '提出調整者']
    ws.append(header)

    for e in rows:
        sid = int(e['id'])
        changed = d.suggesters(sid)
        row = [sid, e.get('column2', ''), e.get('column3', ''), e.get('column4', ''),
               '、'.join(d.live_tags(sid)), d.cell_text(sid, d.baseline)]
        row += [d.cell_text(sid, p['id']) for p in d.others]
        row += [len(changed),
                '、'.join(p['short'] for p in d.others if p['id'] in changed)]
        ws.append(row)

    ws.freeze_panes = 'C2'
    set_widths(ws, [6, 12, 40, 52, 22, 26] + [30] * len(d.others) + [10, 22])
    style_header(ws)

    first_prop = 7
    for r, e in enumerate(rows, start=2):
        sid = int(e['id'])
        changed = set(d.suggesters(sid))
        for cell in ws[r]:
            cell.alignment = TOP_WRAP
            cell.border = BORDER
        ws.cell(row=r, column=5).fill = LIVE_FILL
        ws.cell(row=r, column=6).fill = BASE_FILL
        for k, p in enumerate(d.others):
            if p['id'] in changed:
                ws.cell(row=r, column=first_prop + k).fill = DIFF_FILL
    ws.auto_filter.ref = 'A1:%s%d' % (get_column_letter(ws.max_column), ws.max_row)
    return title_note


def build_tab_sheet(ws, d, tab_name):
    """單一分頁籤：列出被歸入這個分類的大事（依現場操作分類）。"""
    ws.append(['序號', '日期', '大事', '紀要', '本則的現場操作分類',
               '採用此分類者', '定案是否含此分類'])
    n = 0
    for e in d.events:
        sid = int(e['id'])
        live = d.live_tags(sid)
        if tab_name not in live:
            continue
        n += 1
        adopters = [p['short'] if p['id'] != d.baseline else '定案'
                    for p in d.proposals
                    if tab_name in d.canon(d.tags(sid, p['id']))]
        ws.append([sid, e.get('column2', ''), e.get('column3', ''), e.get('column4', ''),
                   '、'.join(live), '、'.join(adopters),
                   '是' if tab_name in d.canon(d.tags(sid, d.baseline)) else '否'])
    ws.freeze_panes = 'C2'
    set_widths(ws, [6, 12, 44, 60, 22, 30, 14])
    style_header(ws, TAB_HEADER_FILL)
    for r in range(2, ws.max_row + 1):
        for cell in ws[r]:
            cell.alignment = TOP_WRAP
            cell.border = BORDER
        ws.cell(row=r, column=5).fill = LIVE_FILL
    ws.auto_filter.ref = 'A1:G%d' % ws.max_row
    return n


def build(out_dir):
    d = Data()
    diff_rows = [e for e in d.events if d.suggesters(int(e['id']))]

    wb = Workbook()
    ws_all = wb.active
    ws_all.title = '總表'
    build_main_sheet(ws_all, d, d.events, '全部')
    build_main_sheet(wb.create_sheet('差異表'), d, diff_rows, '差異')

    counts = []
    for tab in d.tab_order:
        n = build_tab_sheet(wb.create_sheet(tab), d, tab)
        counts.append((tab, n))

    info = wb.create_sheet('說明')
    lines = [
        ['正覺大事紀　分類彙整（含現場操作分類）'],
        [],
        ['總表', '196 則全部'],
        ['差異表', '%d 則：至少有一位菩薩的分類與「1150810 會議分頁籤定案-初版」不同' % len(diff_rows)],
        ['各分頁籤工作表', '每個分頁籤一張，列出該分類收錄了哪些大事，依「現場操作分類」歸類'],
        [],
        ['現場操作分類',
         '最大化綜合：把定案與 8 位菩薩提出過的分頁籤全部列上，依分頁籤原始順序排列。'
         '因此同一則大事可能同時出現在多個分頁籤工作表中。'],
        ['各菩薩欄位', '分類寫在第一行，該人的說明文字照錄於同一儲存格的次行以下，未改寫、未補述'],
        ['大事／日期／紀要', '照錄原始資料，未做任何更動'],
        ['底色', '綠＝定案；淺綠＝現場操作分類；橘＝該案與定案不同'],
        [],
        ['各分頁籤則數'],
    ]
    lines += [['　' + t, '%d 則' % n] for t, n in counts]
    lines += [
        [],
        ['資料來源', '正覺大事紀_1150810會議分頁籤定案.xlsx'],
        ['', '正覺大事紀_總表分類二次作業檔-淳渝.xlsx（淳渝菩薩之分類與說明以此為準）'],
        ['', '大事、日期、紀要取自 table_data.js（同原始大事紀檔）'],
        [],
        ['注意', '其餘 7 位菩薩之分類仍取自 1150810 定案檔中的第一次分類欄位；'
                 '若要改用各自的二次作業檔，需另行更新。'],
    ]
    for line in lines:
        info.append(line)
    info.column_dimensions['A'].width = 20
    info.column_dimensions['B'].width = 82
    info['A1'].font = Font(bold=True, size=14)
    for r in range(3, info.max_row + 1):
        info.cell(row=r, column=1).font = Font(bold=True)
        info.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical='top')

    os.makedirs(out_dir, exist_ok=True)
    path = os.path.join(out_dir, OUT_NAME)
    wb.save(path)
    return path, len(d.events), len(diff_rows), counts


def main():
    out_dir = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_OUT
    path, total, diff, counts = build(out_dir)
    print('已輸出：%s' % path)
    print('總表 %d 則｜差異表 %d 則' % (total, diff))
    for t, n in counts:
        print('  %-6s %3d 則' % (t, n))


if __name__ == '__main__':
    main()
