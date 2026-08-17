# -*- coding: utf-8 -*-
"""從 1150810 會議定案檔產生第二層（主畫面）所需的兩個資料檔。

輸入：
  1. 正覺大事紀_1150810會議分頁籤定案.xlsx（工作表「大事」「分類」）
     - 「1150810會議分頁籤定案-初版」欄 = 定案分類（196 則全部有值）
     - 其後 8 欄 = 各人對定案的調整建議（空白＝同意定案，照錄原文）
  2. 03  出版品總表-書號&書名-20251219.xlsx（561 筆出版品）
  3. table_data.js（196 則大事的日期／大事／紀要，文字一律不動）

輸出：
  meeting2_data.js          分頁籤定義、逐則分類、各人原文說明
  publications_full_data.js 出版品總表（每一集一列）＋對應大事之「出版品紀要」

用法：
  python build_meeting2_data.py [定案xlsx路徑]
"""
import json
import os
import re
import sys

import openpyxl

BASE = os.path.dirname(os.path.abspath(__file__))
DEFAULT_SRC = os.path.join(BASE, '正覺大事紀_1150810會議分頁籤定案.xlsx')
PUB_SRC = os.path.join(BASE, '03  出版品總表-書號&書名-20251219.xlsx')

# 分頁籤（順序照錄定案檔「分類」工作表）
TABS = ['導師弘法', '組織發展', '共修紀事', '摧邪顯正', '公益推廣', '出版流通', '綜合紀事']

# 計票與歸類時視為同一分類（畫面上仍顯示原文）
ALIAS = {
    '平實導師': '導師弘法',
    '平實導師弘法': '導師弘法',
    '破邪顯正': '摧邪顯正',
    '公益流通': '公益推廣',
    '共修記事': '共修紀事',
    '綜合記事': '綜合紀事',
}
NAMES = sorted(set(TABS) | set(ALIAS.keys()), key=len, reverse=True)

# 提案欄位 -> 提案代號／顯示名稱
COLUMNS = [
    ('D0', '1150810會議分頁籤定案-初版', '定案初版'),
    ('M1', '麗薇菩薩提案', '麗薇'),
    ('M2', '月英菩薩提案', '月英'),
    ('M3', '淑滿菩薩提案', '淑滿'),
    ('M4', '淑妙菩薩提案', '淑妙'),
    ('M5', '劍華菩薩提案', '劍華'),
    ('M6', '柄富菩薩提案', '柄富'),
    ('M7', '淳渝二次調整', '淳渝'),
    ('M8', '雅鴦菩薩', '雅鴦'),
]

# 人工覆寫：自動解析會誤讀原意者（僅此一則，理由記錄於此）
# 59 雅鴦：原文為「不放進摧邪顯正中」，語意是維持定案（出版流通），非改列摧邪顯正
MANUAL_TAGS = {
    (59, 'M8'): ['出版流通'],
}


def split_note(raw):
    """把儲存格切成「分類部分」與「說明部分」。"""
    text = raw.replace('\r\n', '\n').strip()
    marks = []
    m = re.search(r'說明\s*[:：]?', text)
    if m:
        marks.append(m.start())
    if '\n' in text:
        marks.append(text.index('\n'))
    if marks:
        i = min(marks)
        return text[:i].strip(), text[i:].strip()
    return text, ''


def parse_cell(raw):
    """解析單一儲存格：回傳 (分類清單, 說明, 是否由敘述文字推得)。"""
    head, note = split_note(raw)
    parens = re.findall(r'[（(][^）)]*[）)]', head)
    head_clean = re.sub(r'[（(][^）)]*[）)]', '，', head)
    if parens:
        note = ('　'.join(parens) + ('\n' if note else '') + note).strip()
    # 「建議分類：X」「建議：X」時，以建議的部分為準
    rec = re.search(r'建議(?:分類)?\s*[:：]\s*([^。\n]*)', head_clean)
    segment = rec.group(1) if rec else head_clean

    hits = []
    for name in NAMES:
        for mm in re.finditer(re.escape(name), segment):
            hits.append((mm.start(), name))
    hits.sort()
    tags, seen = [], set()
    for _, name in hits:
        canon = ALIAS.get(name, name)
        if canon not in seen:
            seen.add(canon)
            tags.append(canon)
    # 頭段超過 18 字者代表是敘述句，標記為自動解析、需人工確認
    auto = len(head) > 18 or bool(rec)
    return tags, note, auto


def load_events():
    src = open(os.path.join(BASE, 'table_data.js'), encoding='utf-8').read()
    return json.loads(src[src.index('['):src.rindex(']') + 1])


def build_classification(src_path):
    wb = openpyxl.load_workbook(src_path, data_only=True)
    ws = wb['大事']
    header = [c.value for c in ws[1]]
    col_of = {}
    for pid, title, _short in COLUMNS:
        if title not in header:
            raise SystemExit('定案檔缺少欄位：%s' % title)
        col_of[pid] = header.index(title)

    categories, notes = {}, {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] in (None, ''):
            continue
        sid = int(row[0])
        base = [ALIAS.get(x.strip(), x.strip())
                for x in re.split(r'[、,，]', str(row[col_of['D0']]).strip()) if x.strip()]
        entry = {'D0': base}
        note_entry = {}
        for pid, title, _short in COLUMNS[1:]:
            raw = row[col_of[pid]]
            raw = '' if raw is None else str(raw).strip()
            if not raw:
                entry[pid] = list(base)          # 空白＝同意定案
                continue
            tags, note, auto = parse_cell(raw)
            if (sid, pid) in MANUAL_TAGS:
                tags, auto = MANUAL_TAGS[(sid, pid)], False
            if not tags:
                tags = list(base)                # 只留意見、未指定分類者維持定案
            entry[pid] = tags
            note_entry[pid] = {'raw': raw, 'auto': auto, 'changed': tags != base}
        categories[str(sid)] = entry
        if note_entry:
            notes[str(sid)] = note_entry
    return categories, notes


# ===== 出版品總表 =====
def norm_title(text):
    t = str(text).replace('\xa0', ' ').strip()
    t = re.sub(r'[（(][^）)]*[）)]', '', t)
    t = re.sub(r'第[一二三四五六七八九十百零]+輯|第\s*\d+\s*輯', '', t)
    t = re.sub(r'[上中下]冊|上下冊|精裝本|平裝本', '', t)
    t = re.sub(r'[《》\s‧·．,，、─—–\-~～]', '', t)
    return t


def stem_title(text):
    return re.sub(r'(講義|講記|詳解|宗通|真義)$', '', text)


# 人工對應（書名與大事用字不同，但同一本書）
MANUAL_PUB = {
    '正覺電子書': 152,
    '現代人應有的宗教觀': 88,
}


def build_publications(events):
    index = {}
    for ev in events:
        eid = int(ev['id'])
        for book in re.findall(r'《([^》]+)》', ev['column3']):
            index.setdefault(norm_title(book), []).append((eid, '大事'))
        for book in re.findall(r'《([^》]+)》', ev['column4']):
            index.setdefault(norm_title(book), []).append((eid, '紀要'))
    ev_by_id = {int(e['id']): e for e in events}

    wb = openpyxl.load_workbook(PUB_SRC, data_only=True)
    ws = wb.worksheets[0]

    def fmt(v):
        if v is None:
            return ''
        if hasattr(v, 'strftime'):
            return v.strftime('%Y/%m/%d')
        return str(v).strip()

    out = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if all(x in (None, '') for x in row[:3]):
            continue
        title = str(row[2] or '').replace('\xa0', ' ').strip()
        if not title:
            continue
        key = norm_title(title)
        hit, how = None, ''
        if title in MANUAL_PUB:
            hit, how = MANUAL_PUB[title], '人工對應'
        if hit is None:
            # 以分數挑最合適的一則大事：書名完全相同 > 書名相含；「大事」欄 > 「紀要」欄；
            # 出版／發行／創刊類大事優先（出版品的來源事件）
            best = None
            for k, refs in index.items():
                if not k:
                    continue
                if k == key:
                    base_score, kind = 120, '書名相同'
                elif (len(k) >= 3 and k in key) or (len(key) >= 3 and key in k):
                    base_score, kind = 40 + len(k), '書名相含'
                else:
                    continue
                for eid, field in refs:
                    score = base_score + (20 if field == '大事' else 0)
                    title_ev = ev_by_id[eid]['column3']
                    if re.search(r'出版|發行|創刊', title_ev):
                        score += 10
                    if best is None or score > best[0]:
                        best = (score, eid, kind if field == '大事' else kind + '（紀要）')
            if best:
                hit, how = best[1], best[2]
        if hit is None:
            s = stem_title(key)
            for k, v in index.items():
                ks = stem_title(k)
                if ks and (ks == s or (len(ks) >= 2 and ks in s) or (len(s) >= 2 and s in ks)):
                    hit, how = v[0][0], '書名主體'
                    break
        item = {
            'bookNo': fmt(row[0]),
            'category': fmt(row[1]),
            'title': title,
            'publishDate': fmt(row[3]),
            'copyrightDate': fmt(row[4]),
            'note': fmt(row[5]) if len(row) > 5 else '',
            'eventId': hit,
            'eventTitle': ev_by_id[hit]['column3'] if hit else '',
            'eventDate': ev_by_id[hit]['column2'] if hit else '',
            'summary': ev_by_id[hit]['column4'] if hit else '',
            'matchBy': how,
        }
        out.append(item)
    return out


def main():
    src_path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_SRC
    events = load_events()
    categories, notes = build_classification(src_path)

    proposals = []
    for pid, title, short in COLUMNS:
        proposals.append({
            'id': pid,
            'name': title if pid == 'D0' else title,
            'short': short,
            'author': short,
            'baseline': pid == 'D0',
            'source': os.path.basename(src_path),
            'intro': ('1150810 會議之分頁籤定案（初版），196 則大事全部已歸類，'
                      '為本頁其餘各案的比較基準。'
                      if pid == 'D0' else
                      '以「1150810 會議分頁籤定案-初版」為底稿，僅列出'
                      + short + '菩薩提出調整建議之則次；其餘則次即表示同意定案。'),
            'tabs': [{'name': t, 'desc': ''} for t in TABS],
        })

    changed = {pid: 0 for pid, _, _ in COLUMNS}
    for sid, entry in categories.items():
        for pid, _t, _s in COLUMNS[1:]:
            if entry[pid] != entry['D0']:
                changed[pid] += 1

    header = (
        '// 1150810 會議分頁籤定案（初版）＋各菩薩調整建議\n'
        '// 資料來源：%s（工作表「大事」「分類」）\n'
        '// 定案欄 196 則全部有值；各人欄空白＝同意定案，故其分類即沿用定案。\n'
        '// 各人欄若為敘述文字，程式以關鍵字解析出分類（eventProposalNotes[].auto=true），\n'
        '// 並一律保留原文（.raw）於畫面上對照，未改寫、未補述任何內容。\n'
        '// 大事、日期、紀要文字一律取自 table_data.js，不做任何更動。\n'
        % os.path.basename(src_path)
    )
    body = [header]
    body.append('const TAB_ORDER = ' + json.dumps(TABS, ensure_ascii=False) + ';\n')
    body.append('const BASELINE_ID = "D0";\n')
    body.append('const proposals = ' + json.dumps(proposals, ensure_ascii=False, indent=2) + ';\n')
    body.append('const tagAliases = ' + json.dumps(ALIAS, ensure_ascii=False, indent=2) + ';\n')
    body.append('// eventProposalCategories[大事序號] = { 提案代號: [分類, ...] }\n')
    body.append('const eventProposalCategories = '
                + json.dumps(categories, ensure_ascii=False, indent=1) + ';\n')
    body.append('// eventProposalNotes[大事序號][提案代號] = { raw: 原文, auto: 是否由敘述解析, changed: 是否與定案不同 }\n')
    body.append('const eventProposalNotes = '
                + json.dumps(notes, ensure_ascii=False, indent=1) + ';\n')
    with open(os.path.join(BASE, 'meeting2_data.js'), 'w', encoding='utf-8') as f:
        f.write('\n'.join(body))

    pubs = build_publications(events)
    matched = sum(1 for p in pubs if p['eventId'])
    with open(os.path.join(BASE, 'publications_full_data.js'), 'w', encoding='utf-8') as f:
        f.write('// 出版品總表（每一集一列，共 %d 筆）\n' % len(pubs))
        f.write('// 資料來源：03  出版品總表-書號&書名-20251219.xlsx\n')
        f.write('// summary（出版品紀要）＝對應大事之「紀要」原文，照錄自 table_data.js，未改寫。\n')
        f.write('// 找不到對應大事者 eventId 為 null，紀要留白（不臆測、不補述）。目前已對應 %d 筆。\n'
                % matched)
        f.write('const publicationsFull = ' + json.dumps(pubs, ensure_ascii=False, indent=1) + ';\n')

    print('大事：%d 則' % len(categories))
    for pid, title, short in COLUMNS[1:]:
        print('  %-14s 提出調整：%3d 則' % (title, changed[pid]))
    print('出版品：%d 筆，已對應大事 %d 筆，未對應 %d 筆'
          % (len(pubs), matched, len(pubs) - matched))


if __name__ == '__main__':
    main()
